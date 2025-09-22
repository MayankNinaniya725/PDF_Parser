"""
Excel Helper module for creating consistent Excel files

This module provides functions to create Excel files with consistent
formatting and structure for PDF data extraction results.
"""
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.conf import settings
from openpyxl.utils.dataframe import dataframe_to_rows

# Define consistent styles
HEADER_FONT = Font(name='Arial', size=12, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

def create_template_if_not_exists():
    """
    Creates a template Excel file if it doesn't exist
    This ensures we always have a template to use
    """
    template_path = os.path.join(settings.BASE_DIR, 'extractor', 'templates', 'excel', 'extraction_template.xlsx')
    
    # Only create if it doesn't exist
    if not os.path.exists(template_path):
        # Create a workbook with the sheets we need
        wb = openpyxl.Workbook()
        
        # Create Summary sheet
        summary_sheet = wb.active
        summary_sheet.title = 'Summary'
        summary_sheet['A1'] = 'Information'
        summary_sheet['B1'] = 'Value'
        
        # Add sample data rows (these will be replaced)
        info_items = ['File Name', 'Vendor', 'Upload Date', 'Total Fields', 'Total Pages', 'Status']
        for idx, item in enumerate(info_items, start=2):
            summary_sheet[f'A{idx}'] = item
            summary_sheet[f'B{idx}'] = ''
        
        # Create Extracted Data sheet
        data_sheet = wb.create_sheet('Extracted Data')
        headers = ['Field Type', 'Extracted Value', 'Page Number', 'PDF Location', 'Extracted At']
        for col_idx, header in enumerate(headers, start=1):
            data_sheet.cell(row=1, column=col_idx, value=header)
        
        # Create Key Fields sheet
        key_sheet = wb.create_sheet('Key Fields')
        key_headers = ['Field', 'Value', 'Page', 'PDF File', 'Status']
        for col_idx, header in enumerate(key_headers, start=1):
            key_sheet.cell(row=1, column=col_idx, value=header)
        
        # Create Page Summary sheet
        page_sheet = wb.create_sheet('Page Summary')
        page_headers = ['Page Number', 'Fields Found', 'PDF File', 'Key Fields Found']
        for col_idx, header in enumerate(page_headers, start=1):
            page_sheet.cell(row=1, column=col_idx, value=header)
        
        # Apply formatting to all sheets
        for sheet in wb.worksheets:
            # Format headers
            for cell in sheet[1]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = BORDER
            
            # Set column widths
            for idx, column in enumerate(sheet.columns, start=1):
                sheet.column_dimensions[get_column_letter(idx)].width = 18
                
            # Last column is usually wider (for comments, long text)
            last_col = get_column_letter(sheet.max_column)
            sheet.column_dimensions[last_col].width = 30
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(template_path), exist_ok=True)
        
        # Save the template
        wb.save(template_path)
    
    return template_path

def apply_formatting(workbook):
    """Apply consistent formatting to all sheets in the workbook"""
    for sheet in workbook.worksheets:
        # Format headers
        for cell in sheet[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER
        
        # Auto-adjust column widths based on content
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = min(adjusted_width, 40)  # Cap at 40 for readability
        
        # Add borders to all data cells
        for row in sheet.iter_rows(min_row=2):  # Skip header
            for cell in row:
                cell.border = BORDER

def write_dataframe_to_sheet(workbook, sheet_name, df):
    """Write a pandas DataFrame to a sheet with consistent formatting"""
    # Get or create the sheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Clear existing data (except header)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
    else:
        sheet = workbook.create_sheet(sheet_name)
    
    # Write headers if sheet is empty
    if sheet.max_row < 1:
        for col_idx, column in enumerate(df.columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=column)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER
    
    # Write data
    for row_idx, row_data in enumerate(df.values, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
    
    return sheet
